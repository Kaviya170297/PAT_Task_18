from behave import *
import openpyxl

from pages.login_page import LoginPage
from pages.dashboard_page import DashboardPage

@when(u'I open the Guvi Homepage')
def open_home(context):
    context.driver.get("https://v2.zenclass.in/login")
    context.login_page = LoginPage(context.driver)
    context.dashboard_page = DashboardPage(context.driver)

@when("I read login data from Excel")
def read_data(context):

    context.workbook = openpyxl.load_workbook(
        r"C:\Users\kaviy\PycharmProjects\PAT_Task_20\testdata\login_data.xlsx"
    )

    context.sheet = context.workbook["Sheet1"]

    context.data = []

    for row_index in range(2, context.sheet.max_row + 1):

        username = context.sheet.cell(
            row=row_index,
            column=1
        ).value

        password = context.sheet.cell(
            row=row_index,
            column=2
        ).value

        expected_result = context.sheet.cell(
            row=row_index,
            column=3
        ).value

        # Stop reading when all three cells are empty
        if username is None and password is None and expected_result is None:
            break

        # Convert None values to empty strings
        username = username or ""
        password = password or ""
        expected_result = expected_result or ""

        context.data.append(
            (
                row_index,
                str(username),
                str(password),
                str(expected_result)
            )
        )

@then("I perform login tests for each row")
def validate_users(context):

    login_url = "https://v2.zenclass.in/login"

    for row_index, username, password, expected_result in context.data:

        print("--------------------------------------")
        print("Testing row:", row_index)
        print("Username:", username)

        actual = "FAIL"

        try:
            # Clear previous login session before every row
            context.driver.delete_all_cookies()

            context.driver.execute_script(
                "window.localStorage.clear();"
            )

            context.driver.execute_script(
                "window.sessionStorage.clear();"
            )

            # Open login page
            context.driver.get(login_url)

            # Enter credentials
            context.login_page.enter_username(username)
            context.login_page.enter_password(password)
            context.login_page.click_login()

            # Check whether dashboard loaded
            dashboard_loaded = (
                context.dashboard_page.verify_dashboard_loaded()
            )

            if dashboard_loaded:

                actual = "PASS"

                print("Login successful:", username)

                # Close popup only if it exists
                try:
                    context.dashboard_page.close_popup()
                except Exception:
                    print("Popup was not displayed")

                # Try normal logout
                try:
                    context.dashboard_page.logout()

                    assert (
                        context.dashboard_page.verify_logout()
                    ), "Logout verification failed"

                    print("Logout successful:", username)

                except Exception as logout_error:
                    print("Normal logout failed:", logout_error)

            else:
                error_msg = (
                    context.login_page.get_error_message()
                )

                if error_msg:
                    print(
                        f"Login failed for {username} | "
                        f"Error: {error_msg}"
                    )
                else:
                    print(
                        f"Login failed for {username} | "
                        "No error message found"
                    )

        except Exception as error:

            print(
                f"Test execution failed for {username}"
            )

            print("Reason:", error)

            try:
                error_msg = (
                    context.login_page.get_error_message()
                )

                if error_msg:
                    print("Error message:", error_msg)

            except Exception:
                pass

        finally:
            # This runs after every Excel row,
            # whether the row passed or failed.

            try:
                context.driver.delete_all_cookies()

                context.driver.execute_script(
                    "window.localStorage.clear();"
                )

                context.driver.execute_script(
                    "window.sessionStorage.clear();"
                )

                context.driver.get(login_url)

                print("Session cleared for next row")

            except Exception as clear_error:
                print(
                    "Could not clear session:",
                    clear_error
                )

        # Convert expected value
        expected_text = str(
            expected_result
        ).strip().lower()

        if expected_text in ["pass", "valid"]:
            expected = "PASS"
        else:
            expected = "FAIL"

        # Compare expected and actual
        if expected == actual:
            result = "Matched"
        else:
            result = "Mismatch"

        print("Expected:", expected)
        print("Actual:", actual)
        print("Comparison:", result)

        # Column 4: actual result
        context.sheet.cell(
            row=row_index,
            column=4
        ).value = actual

        # Column 5: comparison
        context.sheet.cell(
            row=row_index,
            column=5
        ).value = result

    # Save Excel after all rows
    context.workbook.save(
        r"C:\Users\kaviy\PycharmProjects"
        r"\PAT_Task_20\testdata\login_data.xlsx"
    )

    context.workbook.close()

    print("All Excel rows completed")

@then(u'Validate Username and Password input boxes')
def step_validate_input_boxes(context):
    context.login_page.validate_input_boxes()

@then(u'Validate Submit button working or not')
def step_validate_submit_button(context):
    context.login_page.validate_submit_button()
